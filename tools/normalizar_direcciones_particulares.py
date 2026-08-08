from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = Path("BOMSA_MA_DIRECCIONESPARTICULARES_20260720_162801.xlsx")
DEFAULT_OUTPUT = Path("BOMSA_MA_DIRECCIONESPARTICULARES_NORMALIZADO_20260720.xlsx")

REGIONS = {
    1: "Tarapaca",
    2: "Antofagasta",
    3: "Atacama",
    4: "Coquimbo",
    5: "Valparaiso",
    6: "Libertador General Bernardo O'Higgins",
    7: "Maule",
    8: "Biobio",
    9: "La Araucania",
    10: "Los Lagos",
    11: "Aysen del General Carlos Ibanez del Campo",
    12: "Magallanes y de la Antartica Chilena",
    13: "Metropolitana de Santiago",
    14: "Los Rios",
    15: "Arica y Parinacota",
    16: "Nuble",
}

COMMUNE_REFERENCE = {
    "santiago": {"region_code": 13, "region": REGIONS[13], "lat": -33.4489, "lon": -70.6693},
    "puerto octay": {"region_code": 10, "region": REGIONS[10], "lat": -40.9744, "lon": -72.8847},
    "san vicente de tagua": {"region_code": 6, "region": REGIONS[6], "lat": -34.4386, "lon": -71.0777},
    "san vicente de tagua tagua": {"region_code": 6, "region": REGIONS[6], "lat": -34.4386, "lon": -71.0777},
    "caldera": {"region_code": 3, "region": REGIONS[3], "lat": -27.0667, "lon": -70.8167},
    "pichidegua": {"region_code": 6, "region": REGIONS[6], "lat": -34.3583, "lon": -71.2833},
    "villa alemana": {"region_code": 5, "region": REGIONS[5], "lat": -33.0422, "lon": -71.3733},
    "colina": {"region_code": 13, "region": REGIONS[13], "lat": -33.2044, "lon": -70.6756},
    "pucon": {"region_code": 9, "region": REGIONS[9], "lat": -39.2822, "lon": -71.9547},
    "vina del mar": {"region_code": 5, "region": REGIONS[5], "lat": -33.0245, "lon": -71.5518},
    "illapel": {"region_code": 4, "region": REGIONS[4], "lat": -31.6308, "lon": -71.1653},
    "curacavi": {"region_code": 13, "region": REGIONS[13], "lat": -33.3986, "lon": -71.1281},
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize(value: object) -> str:
    text = clean(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.casefold().split())


def only_int(value: str) -> int | None:
    digits = re.sub(r"\D", "", value or "")
    return int(digits) if digits else None


def split_number_and_unit(number_field: str, unit_field: str) -> tuple[str, str]:
    text = " ".join(part for part in [clean(number_field), clean(unit_field)] if part)
    match = re.match(r"^(\d+[A-Za-z]?)\s*(.*)$", text)
    if not match:
        return clean(number_field), clean(unit_field)
    return match.group(1), match.group(2).strip()


def classify_unit(unit: str, extra: str) -> tuple[str, str, str]:
    combined = " ".join(part for part in [unit, extra] if part).strip()
    normalized = normalize(combined)
    block = ""
    depto = unit.strip()
    detalle = extra.strip()

    block_match = re.search(r"\b(block|blok|blk|torre)\s*([a-z0-9]+)", normalized, re.IGNORECASE)
    if block_match:
        block = block_match.group(2).upper()
    depto_match = re.search(r"\b(depto|departamento|dpto|casa)\.?\s*([a-z0-9-]+)", normalized, re.IGNORECASE)
    if depto_match:
        depto = depto_match.group(2).upper()
    return block, depto, detalle


def parse_address(raw: object) -> dict[str, object]:
    text = clean(raw)
    padded = text.ljust(128)
    street = clean(padded[0:25])
    number_field = padded[25:33]
    unit_field = padded[33:41]
    commune_source_code = clean(padded[41:44])
    commune = clean(padded[44:65]).title()
    region_source_code = clean(padded[65:67])
    extra = clean(padded[67:])

    number, unit = split_number_and_unit(number_field, unit_field)
    block, depto, detalle = classify_unit(unit, extra)
    region_code_raw = only_int(region_source_code)
    commune_key = normalize(commune)
    ref = COMMUNE_REFERENCE.get(commune_key)

    region_code = ref["region_code"] if ref else region_code_raw
    region = ref["region"] if ref else REGIONS.get(region_code or 0, "")
    region_status = "ok"
    if ref and region_code_raw and region_code_raw != ref["region_code"]:
        region_status = f"corregido_desde_origen_{region_code_raw}"
    elif not ref:
        region_status = "sin_referencia_comuna"

    coords = ""
    coords_status = "pendiente"
    if ref:
        coords = f'{ref["lat"]},{ref["lon"]}'
        coords_status = "centroide_comunal_aproximado"

    return {
        "calle": street,
        "numero": number,
        "depto_casa": depto,
        "block_torre": block,
        "detalle_direccion": detalle,
        "comuna": commune,
        "codigo_comuna_origen": commune_source_code,
        "region": region,
        "codigo_region": region_code or "",
        "codigo_region_origen": region_source_code,
        "validacion_region": region_status,
        "coordenadas": coords,
        "coordenadas_estado": coords_status,
    }


def build_output(input_path: Path, output_path: Path) -> None:
    source = load_workbook(input_path, read_only=True, data_only=True)
    source_ws = source.active
    headers = [cell.value for cell in source_ws[1]]
    try:
        address_idx = headers.index("DIRECCION")
    except ValueError as exc:
        raise SystemExit("No se encontro columna DIRECCION") from exc

    output = Workbook()
    ws = output.active
    ws.title = "Direcciones normalizadas"

    parsed_headers = [
        "CALLE",
        "NUMERO",
        "DEPTO_CASA",
        "BLOCK_TORRE",
        "DETALLE_DIRECCION",
        "COMUNA",
        "CODIGO_COMUNA_ORIGEN",
        "REGION",
        "CODIGO_REGION",
        "CODIGO_REGION_ORIGEN",
        "VALIDACION_REGION",
        "COORDENADAS",
        "COORDENADAS_ESTADO",
    ]
    ws.append(headers + parsed_headers)

    for row in source_ws.iter_rows(min_row=2, values_only=True):
        parsed = parse_address(row[address_idx])
        ws.append(
            list(row)
            + [
                parsed["calle"],
                parsed["numero"],
                parsed["depto_casa"],
                parsed["block_torre"],
                parsed["detalle_direccion"],
                parsed["comuna"],
                parsed["codigo_comuna_origen"],
                parsed["region"],
                parsed["codigo_region"],
                parsed["codigo_region_origen"],
                parsed["validacion_region"],
                parsed["coordenadas"],
                parsed["coordenadas_estado"],
            ]
        )

    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = max(len(clean(cell.value)) for cell in column_cells[:200])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)
    ws.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normaliza direcciones particulares y agrega region/coordenadas comunales.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_output(args.input, args.output)
    print(f"Excel generado: {args.output}")


if __name__ == "__main__":
    main()
