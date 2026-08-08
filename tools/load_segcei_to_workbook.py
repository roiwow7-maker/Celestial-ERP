from __future__ import annotations

import argparse
import csv
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


CSV_SEPARATOR = ";"
TARGET_ITEM = "SEGCEI"
TARGET_HEADER = "Seguro Cesantia Empleador (Ley proteccion empleo)"
DOCUMENT_HEADER = "Numero de Documento"
FICHA_HEADER = "Codigo de Ficha"


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("*", "")
    return " ".join(text.casefold().split())


def normalize_code(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def read_segcei_values(csv_path: Path) -> dict[tuple[str, str], int]:
    values: dict[tuple[str, str], int] = defaultdict(int)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=CSV_SEPARATOR)
        required = {"periodo", "codigo", "codigo_item", "monto"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise SystemExit(f"Faltan columnas en CSV: {', '.join(sorted(missing))}")
        for row in reader:
            if normalize_code(row.get("codigo_item")) != TARGET_ITEM:
                continue
            period = str(row.get("periodo") or "").strip()
            code = normalize_code(row.get("codigo"))
            amount_text = str(row.get("monto") or "0").strip().replace(",", ".")
            if not period or not code:
                continue
            values[(period, code)] += int(round(float(amount_text)))
    return dict(values)


def header_map(worksheet, row_idx: int) -> dict[str, int]:
    headers = {}
    for cell in worksheet[row_idx]:
        label = normalize_text(cell.value)
        if label:
            headers[label] = cell.column
    return headers


def find_sheet_and_columns(workbook):
    target_label = normalize_text(TARGET_HEADER)
    document_label = normalize_text(DOCUMENT_HEADER)
    ficha_label = normalize_text(FICHA_HEADER)
    for worksheet in workbook.worksheets:
        for row_idx in range(1, min(10, worksheet.max_row) + 1):
            headers = header_map(worksheet, row_idx)
            target_col = headers.get(target_label)
            document_col = headers.get(document_label)
            ficha_col = headers.get(ficha_label)
            if target_col and document_col and ficha_col:
                return worksheet, row_idx, document_col, ficha_col, target_col
    available = {
        worksheet.title: [cell.value for cell in worksheet[1] if cell.value]
        for worksheet in workbook.worksheets
    }
    raise SystemExit(f"No se encontro hoja con columnas requeridas. Encabezados disponibles: {available}")


def key_from_document(document_value: object, ficha_value: object) -> tuple[str, str] | None:
    document = str(document_value or "").strip()
    ficha = normalize_code(ficha_value)
    if not document or not ficha:
        return None
    if "-" in document:
        period = document.split("-", 1)[0].strip()
    else:
        period = document.split(".", 1)[0].strip()
    if len(period) != 6 or not period.isdigit():
        return None
    return period, ficha


def update_workbook(csv_path: Path, workbook_path: Path, dry_run: bool) -> dict[str, object]:
    values = read_segcei_values(csv_path)
    workbook = load_workbook(workbook_path)
    worksheet, header_row, document_col, ficha_col, target_col = find_sheet_and_columns(workbook)

    matched = 0
    changed = 0
    missing_rows = 0
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        key = key_from_document(worksheet.cell(row=row_idx, column=document_col).value, worksheet.cell(row=row_idx, column=ficha_col).value)
        if key is None:
            missing_rows += 1
            continue
        if key not in values:
            continue
        matched += 1
        current_value = worksheet.cell(row=row_idx, column=target_col).value
        new_value = values[key]
        if current_value != new_value:
            changed += 1
            if not dry_run:
                worksheet.cell(row=row_idx, column=target_col).value = new_value

    backup_path = None
    if not dry_run:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = workbook_path.with_name(f"{workbook_path.stem}.backup_segcei_{timestamp}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)
        workbook.save(workbook_path)

    unmatched_values = len(values) - matched
    return {
        "csv_segcei_keys": len(values),
        "sheet": worksheet.title,
        "header_row": header_row,
        "matched_rows": matched,
        "changed_rows": changed,
        "rows_without_key": missing_rows,
        "csv_keys_without_sheet_row": unmatched_values,
        "backup": str(backup_path) if backup_path else "",
        "dry_run": dry_run,
    }


def verify_workbook(csv_path: Path, workbook_path: Path) -> dict[str, object]:
    values = read_segcei_values(csv_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet, header_row, document_col, ficha_col, target_col = find_sheet_and_columns(workbook)
    checked = 0
    mismatches = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        key = key_from_document(worksheet.cell(row=row_idx, column=document_col).value, worksheet.cell(row=row_idx, column=ficha_col).value)
        if key is None or key not in values:
            continue
        checked += 1
        actual = worksheet.cell(row=row_idx, column=target_col).value
        expected = values[key]
        if int(actual or 0) != expected:
            mismatches.append((row_idx, key[0], key[1], expected, actual))
            if len(mismatches) >= 10:
                break
    return {
        "sheet": worksheet.title,
        "checked_rows": checked,
        "mismatch_count_sampled": len(mismatches),
        "sample_mismatches": mismatches,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Carga SEGCEI a Liquidaciones_Historicas_Cargadas.xlsx.")
    parser.add_argument("--csv", type=Path, required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--verify", action="store_true")
    args = parser.parse_args()

    if args.verify:
        result = verify_workbook(args.csv, args.workbook)
    else:
        result = update_workbook(args.csv, args.workbook, args.dry_run)
    for key, value in result.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
