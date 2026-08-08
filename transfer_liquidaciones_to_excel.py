from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable

import openpyxl


DEFAULT_TEMPLATE = Path("Copia de Liquidaciones Históricas (37).xlsx")
DEFAULT_INPUT_DIR = Path("csv_equivalentes_liquidaciones")
DEFAULT_OUTPUT = Path("Liquidaciones_Historicas_Cargadas.xlsx")
CSV_SEPARATOR = ";"

SHEET_TO_CSV = {
    "Liquidaciones": "Liquidaciones.csv",
    "Haberes Imponibles": "Haberes_Imponibles.csv",
    "Haberes No Imponibles": "Haberes_No_Imponibles.csv",
    "Descuentos": "Descuentos.csv",
    "Líneas de Finiquito": "Líneas_de_Finiquito.csv",
}

TEXT_COLUMNS = {
    "Número de Documento*",
    "Código de Ficha",
    "RUT Empresa*",
    "Rut empresa",
    "Nombre",
    "Codigo item",
    "Tributable",
}


def clean_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def sheet_headers(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    return [
        clean_header(cell.value)
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        if clean_header(cell.value)
    ]


def read_csv_headers(csv_path: Path) -> list[str]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=CSV_SEPARATOR)
        return next(reader)


def validate_headers(sheet_name: str, excel_headers: list[str], csv_headers: list[str]) -> None:
    if excel_headers != csv_headers:
        excel_only = [header for header in excel_headers if header not in csv_headers]
        csv_only = [header for header in csv_headers if header not in excel_headers]
        raise ValueError(
            f"Los encabezados no coinciden para la hoja {sheet_name}.\n"
            f"Solo en Excel: {excel_only}\n"
            f"Solo en CSV: {csv_only}"
        )


def clear_sheet_data(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)


def convert_cell_value(header: str, value: str) -> object:
    value = value.strip()
    if value == "":
        return None
    if header in TEXT_COLUMNS:
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return value


def csv_rows(csv_path: Path, headers: list[str]) -> Iterable[list[object]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=CSV_SEPARATOR)
        for row in reader:
            yield [convert_cell_value(header, row.get(header, "")) for header in headers]


def transfer_sheet(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    csv_path: Path,
) -> int:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja en la plantilla: {sheet_name}")
    if not csv_path.exists():
        raise FileNotFoundError(f"No existe el CSV requerido: {csv_path}")

    worksheet = workbook[sheet_name]
    headers = sheet_headers(worksheet)
    validate_headers(sheet_name, headers, read_csv_headers(csv_path))
    clear_sheet_data(worksheet)

    row_count = 0
    for row in csv_rows(csv_path, headers):
        worksheet.append(row)
        row_count += 1
    return row_count


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Traspasa los CSV equivalentes a la plantilla Excel de liquidaciones."
    )
    parser.add_argument(
        "-t",
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help=f"Excel plantilla. Default: {DEFAULT_TEMPLATE}",
    )
    parser.add_argument(
        "-i",
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"Carpeta con CSV equivalentes. Default: {DEFAULT_INPUT_DIR}",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Excel generado. Default: {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook = openpyxl.load_workbook(args.template)

    summary_rows = []
    for sheet_name, csv_name in SHEET_TO_CSV.items():
        csv_path = args.input_dir / csv_name
        rows = transfer_sheet(workbook, sheet_name, csv_path)
        summary_rows.append((sheet_name, csv_path, rows))

    workbook.save(args.output)

    print(f"Excel generado: {args.output}")
    for sheet_name, csv_path, rows in summary_rows:
        print(f"{sheet_name}: {rows} filas desde {csv_path}")


if __name__ == "__main__":
    main()
