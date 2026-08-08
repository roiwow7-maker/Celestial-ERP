from __future__ import annotations

import argparse
import csv
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


CSV_SEPARATOR = ";"
SOURCE_CODE_HEADER = "codigo"
SOURCE_PERIOD_HEADER = "periodo"
DOCUMENT_HEADER = "Numero de Documento"
FICHA_HEADER = "Codigo de Ficha"

DEFAULT_SOURCE = Path("ITEMS_ACUMULADOS_Historico Payroll.xlsx")
DEFAULT_WORKBOOK = Path("Liquidaciones_Historicas_Cargadas.xlsx")

ITEM_TARGETS = {
    "COSESO": "Cotizacion Expectativa de Vida",
    "SEGCEI": "Seguro Cesantia Empleador (Ley proteccion empleo)",
}


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("*", "")
    return " ".join(text.casefold().split())


def normalize_code(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def normalize_period(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def amount_to_int(value: object) -> int:
    if value is None:
        return 0
    text = str(value).strip().replace(",", ".")
    if not text:
        return 0
    return int(round(float(text)))


def source_columns(source_path: Path, item_code: str) -> dict[str, str]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    by_normalized = {normalize_text(header): header for header in headers}
    required = {
        "codigo": SOURCE_CODE_HEADER,
        "periodo": SOURCE_PERIOD_HEADER,
        item_code.casefold(): item_code,
    }
    resolved = {}
    for normalized, logical_name in required.items():
        if normalized not in by_normalized:
            raise SystemExit(f"Falta columna {logical_name!r} en {source_path}")
        resolved[logical_name] = by_normalized[normalized]
    return resolved


def read_item_values(source_path: Path, item_code: str) -> dict[tuple[str, str], int]:
    if source_path.suffix.lower() == ".csv":
        return read_item_values_from_csv(source_path, item_code)

    columns = source_columns(source_path, item_code)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_to_index = {cell.value: index for index, cell in enumerate(worksheet[1])}
    period_idx = header_to_index[columns[SOURCE_PERIOD_HEADER]]
    code_idx = header_to_index[columns[SOURCE_CODE_HEADER]]
    item_idx = header_to_index[columns[item_code]]
    values: dict[tuple[str, str], int] = defaultdict(int)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        period = normalize_period(row[period_idx])
        code = normalize_code(row[code_idx])
        amount = amount_to_int(row[item_idx])
        if not period or not code or amount == 0:
            continue
        values[(period, code)] += amount
    return dict(values)


def read_item_values_from_csv(source_path: Path, item_code: str) -> dict[tuple[str, str], int]:
    values: dict[tuple[str, str], int] = defaultdict(int)
    with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=CSV_SEPARATOR)
        required = {"periodo", "codigo", "codigo_item", "monto"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise SystemExit(f"Faltan columnas en CSV: {', '.join(sorted(missing))}")
        for row in reader:
            if normalize_code(row.get("codigo_item")) != item_code:
                continue
            period = normalize_period(row.get("periodo"))
            code = normalize_code(row.get("codigo"))
            amount = amount_to_int(row.get("monto"))
            if not period or not code or amount == 0:
                continue
            values[(period, code)] += amount
    return dict(values)


def header_map(worksheet, row_idx: int) -> dict[str, int]:
    headers = {}
    for cell in worksheet[row_idx]:
        label = normalize_text(cell.value)
        if label:
            headers[label] = cell.column
    return headers


def find_sheet_and_columns(workbook, target_header: str):
    target_label = normalize_text(target_header)
    document_label = normalize_text(DOCUMENT_HEADER)
    ficha_label = normalize_text(FICHA_HEADER)
    for worksheet in workbook.worksheets:
        for row_idx in range(1, min(10, worksheet.max_row) + 1):
            headers = header_map(worksheet, row_idx)
            target_col = headers.get(target_label)
            document_col = headers.get(document_label)
            ficha_col = headers.get(ficha_label)
            if target_col and document_col and ficha_col:
                return worksheet, row_idx, document_col, ficha_col, target_col
    raise SystemExit(f"No se encontro hoja con columnas requeridas para {target_header!r}.")


def key_from_document(document_value: object, ficha_value: object) -> tuple[str, str] | None:
    document = normalize_period(document_value)
    ficha = normalize_code(ficha_value)
    if not document or not ficha:
        return None
    period = document.split("-", 1)[0].split(".", 1)[0].strip()
    if len(period) != 6 or not period.isdigit():
        return None
    return period, ficha


def update_workbook(
    source_path: Path,
    workbook_path: Path,
    item_code: str,
    target_header: str,
    output_path: Path | None,
    dry_run: bool,
) -> dict[str, object]:
    values = read_item_values(source_path, item_code)
    workbook = load_workbook(workbook_path)
    worksheet, header_row, document_col, ficha_col, target_col = find_sheet_and_columns(workbook, target_header)

    matched = 0
    changed = 0
    rows_without_key = 0
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        key = key_from_document(
            worksheet.cell(row=row_idx, column=document_col).value,
            worksheet.cell(row=row_idx, column=ficha_col).value,
        )
        if key is None:
            rows_without_key += 1
            continue
        if key not in values:
            continue
        matched += 1
        new_value = values[key]
        cell = worksheet.cell(row=row_idx, column=target_col)
        if amount_to_int(cell.value) != new_value:
            changed += 1
            if not dry_run:
                cell.value = new_value

    backup_path = None
    saved_to = ""
    if not dry_run:
        destination = output_path or workbook_path
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination == workbook_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = workbook_path.with_name(f"{workbook_path.stem}.backup_{item_code.lower()}_{timestamp}{workbook_path.suffix}")
            shutil.copy2(workbook_path, backup_path)
        workbook.save(destination)
        saved_to = str(destination)

    return {
        "item_code": item_code,
        "target_header": target_header,
        "source_keys": len(values),
        "sheet": worksheet.title,
        "header_row": header_row,
        "matched_rows": matched,
        "changed_rows": changed,
        "rows_without_key": rows_without_key,
        "source_keys_without_sheet_row": len(values) - matched,
        "backup": str(backup_path) if backup_path else "",
        "saved_to": saved_to,
        "dry_run": dry_run,
    }


def verify_workbook(source_path: Path, workbook_path: Path, item_code: str, target_header: str) -> dict[str, object]:
    values = read_item_values(source_path, item_code)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet, header_row, document_col, ficha_col, target_col = find_sheet_and_columns(workbook, target_header)
    checked = 0
    mismatches = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        key = key_from_document(
            worksheet.cell(row=row_idx, column=document_col).value,
            worksheet.cell(row=row_idx, column=ficha_col).value,
        )
        if key is None or key not in values:
            continue
        checked += 1
        expected = values[key]
        actual = amount_to_int(worksheet.cell(row=row_idx, column=target_col).value)
        if actual != expected:
            mismatches.append((row_idx, key[0], key[1], expected, actual))
            if len(mismatches) >= 10:
                break
    return {
        "item_code": item_code,
        "target_header": target_header,
        "sheet": worksheet.title,
        "checked_rows": checked,
        "mismatch_count_sampled": len(mismatches),
        "sample_mismatches": mismatches,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Carga un item historico a una columna de Liquidaciones.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--item-code", default="COSESO")
    parser.add_argument("--target-header", default="")
    parser.add_argument("--output", type=Path, default=None, help="Si se indica, guarda en otra ruta en vez de sobrescribir.")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--verify", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    item_code = normalize_code(args.item_code)
    target_header = args.target_header or ITEM_TARGETS.get(item_code)
    if not target_header:
        raise SystemExit(f"Indica --target-header para el item {item_code}")

    if args.verify:
        result = verify_workbook(args.source, args.workbook, item_code, target_header)
    else:
        result = update_workbook(
            args.source,
            args.workbook,
            item_code,
            target_header,
            args.output,
            args.dry_run,
        )
    for key, value in result.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
